from django.shortcuts import render
from attendance.models import AttendanceSummary, AttendanceLog
from master.models import Employee, Department
from leave.models import LeaveApplication
from django.db.models import Count, Q
from datetime import date, timedelta
import csv
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill
from django.contrib.auth.decorators import login_required

@login_required
def reports_view(request):
    """Reports page with various attendance and leave reports"""
    today = date.today()
    start_of_month = today.replace(day=1)
    start_of_year = today.replace(month=1, day=1)

    # Monthly attendance report
    monthly_report = AttendanceSummary.objects.filter(
        date__gte=start_of_month,
        date__lte=today
    ).values('employee__department__department_name').annotate(
        total_days=Count('summary_id'),
        present_days=Count('summary_id', filter=Q(status='present')),
        absent_days=Count('summary_id', filter=Q(status='absent')),
        late_days=Count('summary_id', filter=Q(late_by__gt=0))
    )

    # Calculate attendance percentage for monthly report
    for dept in monthly_report:
        if dept['total_days'] > 0:
            dept['attendance_percentage'] = round((dept['present_days'] / dept['total_days']) * 100, 1)
        else:
            dept['attendance_percentage'] = 0

    # Employee-wise report
    employee_report = AttendanceSummary.objects.filter(
        date__gte=start_of_month,
        date__lte=today
    ).values('employee__employee_code', 'employee__first_name', 'employee__last_name').annotate(
        total_days=Count('summary_id'),
        present_days=Count('summary_id', filter=Q(status='present')),
        absent_days=Count('summary_id', filter=Q(status='absent')),
        late_days=Count('summary_id', filter=Q(late_by__gt=0))
    )

    # Calculate attendance percentage for employee report
    for emp in employee_report:
        if emp['total_days'] > 0:
            emp['attendance_percentage'] = round((emp['present_days'] / emp['total_days']) * 100, 1)
        else:
            emp['attendance_percentage'] = 0

    # Leave report
    leave_report = LeaveApplication.objects.filter(
        applied_date__year=today.year
    ).values('leave_type__type_name').annotate(
        total_applications=Count('leave_id'),
        approved=Count('leave_id', filter=Q(status='approved')),
        pending=Count('leave_id', filter=Q(status='pending')),
        rejected=Count('leave_id', filter=Q(status='rejected'))
    )

    # Calculate approval percentage for leave report
    for leave in leave_report:
        if leave['total_applications'] > 0:
            leave['approval_percentage'] = round((leave['approved'] / leave['total_applications']) * 100, 1)
        else:
            leave['approval_percentage'] = 0

    # Device usage report
    device_report = AttendanceLog.objects.filter(
        punch_time__date__gte=start_of_month,
        punch_time__date__lte=today
    ).values('device__device_name').annotate(
        total_punches=Count('attendance_id'),
        in_punches=Count('attendance_id', filter=Q(punch_type='IN')),
        out_punches=Count('attendance_id', filter=Q(punch_type='OUT'))
    )

    context = {
        'user_role': 'Admin',
        'monthly_report': monthly_report,
        'employee_report': employee_report,
        'leave_report': leave_report,
        'device_report': device_report,
        'current_month': today.strftime('%B %Y'),
    }
    return render(request, 'reports/reports.html', context)

@login_required
def export_excel(request):
    """Export reports to Excel"""
    today = date.today()
    start_of_month = today.replace(day=1)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header
    ws['A1'] = f"Attendance Report - {today.strftime('%B %Y')}"
    ws['A1'].font = Font(bold=True, size=14)

    # Department-wise report
    ws['A3'] = "Department-wise Attendance Report"
    ws['A3'].font = Font(bold=True)
    headers = ['Department', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header).font = Font(bold=True)

    monthly_report = AttendanceSummary.objects.filter(
        date__gte=start_of_month,
        date__lte=today
    ).values('employee__department__department_name').annotate(
        total_days=Count('summary_id'),
        present_days=Count('summary_id', filter=Q(status='present')),
        absent_days=Count('summary_id', filter=Q(status='absent')),
        late_days=Count('summary_id', filter=Q(late_by__gt=0))
    )

    row = 5
    for dept in monthly_report:
        attendance_pct = round((dept['present_days'] / dept['total_days']) * 100, 1) if dept['total_days'] > 0 else 0
        ws.cell(row=row, column=1, value=dept['employee__department__department_name'])
        ws.cell(row=row, column=2, value=dept['total_days'])
        ws.cell(row=row, column=3, value=dept['present_days'])
        ws.cell(row=row, column=4, value=dept['absent_days'])
        ws.cell(row=row, column=5, value=dept['late_days'])
        ws.cell(row=row, column=6, value=f"{attendance_pct}%")
        row += 1

    # Employee-wise report
    row += 2
    ws.cell(row=row, column=1, value="Employee-wise Attendance Report").font = Font(bold=True)
    headers = ['Employee', 'Total', 'Present', 'Absent', 'Late', 'Attendance %']
    for col, header in enumerate(headers, 1):
        ws.cell(row=row+1, column=col, value=header).font = Font(bold=True)

    employee_report = AttendanceSummary.objects.filter(
        date__gte=start_of_month,
        date__lte=today
    ).values('employee__employee_code', 'employee__first_name', 'employee__last_name').annotate(
        total_days=Count('summary_id'),
        present_days=Count('summary_id', filter=Q(status='present')),
        absent_days=Count('summary_id', filter=Q(status='absent')),
        late_days=Count('summary_id', filter=Q(late_by__gt=0))
    )

    row += 2
    for emp in employee_report:
        attendance_pct = round((emp['present_days'] / emp['total_days']) * 100, 1) if emp['total_days'] > 0 else 0
        ws.cell(row=row, column=1, value=f"{emp['employee__employee_code']} - {emp['employee__first_name']}")
        ws.cell(row=row, column=2, value=emp['total_days'])
        ws.cell(row=row, column=3, value=emp['present_days'])
        ws.cell(row=row, column=4, value=emp['absent_days'])
        ws.cell(row=row, column=5, value=emp['late_days'])
        ws.cell(row=row, column=6, value=f"{attendance_pct}%")
        row += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{today.strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

@login_required
def export_pdf(request):
    """Export reports to PDF"""
    today = date.today()
    start_of_month = today.replace(day=1)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{today.strftime("%Y%m%d")}.pdf'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title = Paragraph(f"Attendance Report - {today.strftime('%B %Y')}", styles['Title'])
    elements.append(title)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))

    # Department-wise report
    dept_data = [['Department', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %']]
    monthly_report = AttendanceSummary.objects.filter(
        date__gte=start_of_month,
        date__lte=today
    ).values('employee__department__department_name').annotate(
        total_days=Count('summary_id'),
        present_days=Count('summary_id', filter=Q(status='present')),
        absent_days=Count('summary_id', filter=Q(status='absent')),
        late_days=Count('summary_id', filter=Q(late_by__gt=0))
    )

    for dept in monthly_report:
        attendance_pct = f"{round((dept['present_days'] / dept['total_days']) * 100, 1)}%" if dept['total_days'] > 0 else "0%"
        dept_data.append([
            dept['employee__department__department_name'],
            str(dept['total_days']),
            str(dept['present_days']),
            str(dept['absent_days']),
            str(dept['late_days']),
            attendance_pct
        ])

    dept_table = Table(dept_data)
    dept_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(dept_table)
    elements.append(Paragraph("<br/><br/>", styles['Normal']))

    doc.build(elements)
    return response

@login_required
def export_csv(request):
    """Export reports to CSV"""
    today = date.today()
    start_of_month = today.replace(day=1)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename=attendance_report_{today.strftime("%Y%m%d")}.csv'

    writer = csv.writer(response)
    writer.writerow([f'Attendance Report - {today.strftime("%B %Y")}'])
    writer.writerow([])

    # Department-wise report
    writer.writerow(['Department-wise Attendance Report'])
    writer.writerow(['Department', 'Total Days', 'Present', 'Absent', 'Late', 'Attendance %'])

    monthly_report = AttendanceSummary.objects.filter(
        date__gte=start_of_month,
        date__lte=today
    ).values('employee__department__department_name').annotate(
        total_days=Count('summary_id'),
        present_days=Count('summary_id', filter=Q(status='present')),
        absent_days=Count('summary_id', filter=Q(status='absent')),
        late_days=Count('summary_id', filter=Q(late_by__gt=0))
    )

    for dept in monthly_report:
        attendance_pct = f"{round((dept['present_days'] / dept['total_days']) * 100, 1)}%" if dept['total_days'] > 0 else "0%"
        writer.writerow([
            dept['employee__department__department_name'],
            dept['total_days'],
            dept['present_days'],
            dept['absent_days'],
            dept['late_days'],
            attendance_pct
        ])

    writer.writerow([])
    writer.writerow(['Employee-wise Attendance Report'])
    writer.writerow(['Employee', 'Total', 'Present', 'Absent', 'Late', 'Attendance %'])

    employee_report = AttendanceSummary.objects.filter(
        date__gte=start_of_month,
        date__lte=today
    ).values('employee__employee_code', 'employee__first_name', 'employee__last_name').annotate(
        total_days=Count('summary_id'),
        present_days=Count('summary_id', filter=Q(status='present')),
        absent_days=Count('summary_id', filter=Q(status='absent')),
        late_days=Count('summary_id', filter=Q(late_by__gt=0))
    )

    for emp in employee_report:
        attendance_pct = f"{round((emp['present_days'] / emp['total_days']) * 100, 1)}%" if emp['total_days'] > 0 else "0%"
        writer.writerow([
            f"{emp['employee__employee_code']} - {emp['employee__first_name']}",
            emp['total_days'],
            emp['present_days'],
            emp['absent_days'],
            emp['late_days'],
            attendance_pct
        ])

    return response

@login_required
def print_report(request):
    """Print report functionality - returns the same page for printing"""
    return reports_view(request)
